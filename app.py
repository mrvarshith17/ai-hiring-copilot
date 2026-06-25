import streamlit as st
import pandas as pd
import numpy as np
import yaml
import json
import io
import os
from pathlib import Path
from typing import Dict, List, Tuple, Optional
import tempfile
from docx import Document
from PyPDF2 import PdfReader
from sentence_transformers import SentenceTransformer, CrossEncoder, util
import torch
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings

warnings.filterwarnings("ignore")

# Set environment variable for max upload size (2GB = 2048 MB)
os.environ['STREAMLIT_CLIENT_MAX_UPLOAD_SIZE'] = '2048'

# ============================================================================
# PAGE CONFIGURATION & THEME
# ============================================================================
st.set_page_config(
    page_title="SMASHERS AI Hiring Copilot",
    page_icon="🚀",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for dark/professional theme
st.markdown("""
<style>
    :root {
        --primary-bg: #0E1117;
        --secondary-bg: #161B22;
        --accent-color: #58A6FF;
        --text-color: #C9D1D9;
        --success-color: #3FB950;
    }
    
    body {
        color: var(--text-color);
        background-color: var(--primary-bg);
    }
    
    .main {
        background-color: var(--primary-bg);
    }
    
    .stTabs [data-baseweb="tab-list"] button {
        color: var(--text-color);
    }
    
    h1, h2, h3 {
        color: var(--accent-color);
    }
    
    .metric-card {
        background-color: var(--secondary-bg);
        padding: 16px;
        border-radius: 8px;
        border-left: 4px solid var(--accent-color);
    }
</style>
""", unsafe_allow_html=True)

# ============================================================================
# DEVICE & MODEL INITIALIZATION
# ============================================================================
@st.cache_resource
def get_device():
    """Detect GPU availability and return appropriate device."""
    if torch.cuda.is_available():
        return torch.device("cuda")
    else:
        return torch.device("cpu")

@st.cache_resource
def load_semantic_model(model_name: str):
    """Load semantic embedding model."""
    return SentenceTransformer(model_name)

@st.cache_resource
def load_ranking_model(model_name: str):
    """Load cross-encoder ranking model."""
    return CrossEncoder(model_name)

# ============================================================================
# CONFIGURATION LOADING
# ============================================================================
def load_config(config_path: str = "config.yaml") -> Dict:
    """Load configuration from YAML file."""
    try:
        if os.path.exists(config_path):
            with open(config_path, "r") as f:
                return yaml.safe_load(f) or {}
        else:
            return {
                "model_semantic": "sentence-transformers/all-MiniLM-L6-v2",
                "model_ranking": "sentence-transformers/cross-encoder/ms-marco-MiniLM-L6-v2",
                "similarity_threshold": 0.3,
                "semantic_weight": 0.3,
                "ranking_weight": 0.7,
                "top_candidates": 100,
            }
    except Exception as e:
        st.error(f"Error loading config: {e}")
        return {}

# ============================================================================
# FILE PARSING
# ============================================================================
def parse_job_description(file) -> str:
    """Parse DOCX or PDF file and extract text."""
    try:
        if file.name.endswith('.docx'):
            doc = Document(file)
            full_text = "\n".join([para.text for para in doc.paragraphs if para.text.strip()])
        elif file.name.endswith('.pdf'):
            pdf_reader = PdfReader(file)
            full_text = ""
            for page in pdf_reader.pages:
                page_text = page.extract_text()
                if page_text.strip():
                    full_text += page_text + "\n"
        else:
            raise ValueError("Unsupported file format. Please upload .docx or .pdf")

        if not full_text.strip():
            raise ValueError("Job description document is empty.")
        return full_text
    except Exception as e:
        raise Exception(f"Error parsing Job Description: {str(e)}")

def parse_candidate_pool(jsonl_file) -> pd.DataFrame:
    """Parse JSONL file and convert to DataFrame."""
    try:
        candidates = []
        for line in jsonl_file:
            if isinstance(line, bytes):
                line = line.decode("utf-8")
            line = line.strip()
            if line:
                candidate = json.loads(line)
                candidates.append(candidate)
        
        if not candidates:
            raise ValueError("No valid candidate records found in JSONL file.")
        
        return pd.DataFrame(candidates)
    except json.JSONDecodeError as e:
        raise Exception(f"Invalid JSON in candidate file: {str(e)}")
    except Exception as e:
        raise Exception(f"Error parsing candidate pool: {str(e)}")

# ============================================================================
# PIPELINE EXECUTION
# ============================================================================
def stage_1_semantic_filtering(
    job_description: str,
    candidates_df: pd.DataFrame,
    semantic_model: SentenceTransformer,
    similarity_threshold: float,
    top_k: int,
    status_container
) -> Tuple[pd.DataFrame, Dict]:
    """Stage 1: Semantic filtering using embeddings and cosine similarity."""
    with status_container.container():
        st.info("🔍 Generating vector embeddings for Job Description...")
    
    try:
        device = get_device()
        jd_embedding = semantic_model.encode(job_description, convert_to_tensor=True, device=device)
        
        with status_container.container():
            st.info("🔍 Generating embeddings for all candidates (this may take a moment)...")
        
        candidate_summaries = candidates_df["profile"].apply(
            lambda x: x.get("summary", "") if isinstance(x, dict) else ""
        )
        candidate_summaries = candidate_summaries.fillna("")

        candidate_embeddings = semantic_model.encode(
            candidate_summaries.tolist(),
            convert_to_tensor=True,
            device=device,
            show_progress_bar=False
        )
        
        with status_container.container():
            st.info("📊 Computing cosine similarities...")
        
        cos_scores = util.cos_sim(jd_embedding, candidate_embeddings)[0]
        cos_scores_np = cos_scores.cpu().numpy() if hasattr(cos_scores, 'cpu') else cos_scores
        
        candidates_df["semantic_similarity"] = cos_scores_np.flatten()
        
        filtered_candidates = candidates_df[
            candidates_df["semantic_similarity"] >= similarity_threshold
        ].copy()
        
        filtered_candidates = filtered_candidates.sort_values(
            "semantic_similarity",
            ascending=False
        ).head(top_k)
        
        with status_container.container():
            st.success(f"✅ Stage 1 Complete: {len(filtered_candidates)} candidates passed semantic filtering.")
        
        return filtered_candidates, {
            "jd_embedding": jd_embedding,
            "semantic_model": semantic_model
        }
    
    except Exception as e:
        st.error(f"Error in Stage 1 Semantic Filtering: {str(e)}")
        raise

def stage_2_contextual_reranking(
    job_description: str,
    candidates_df: pd.DataFrame,
    ranking_model: CrossEncoder,
    status_container
) -> pd.DataFrame:
    """Stage 2: Contextual re-ranking using cross-encoder."""
    with status_container.container():
        st.info("🎯 Re-ranking candidates using Cross-Encoder (contextual scoring)...")
    
    try:
        candidate_summaries = candidates_df["profile"].apply(
            lambda x: x.get("summary", "") if isinstance(x, dict) else ""
        )
        candidate_summaries = candidate_summaries.fillna("")

        candidate_pairs = [
            [job_description, summary] for summary in candidate_summaries
        ]
        
        scores = ranking_model.predict(candidate_pairs)
        candidates_df["cross_encoder_score"] = scores
        
        with status_container.container():
            st.success(f"✅ Stage 2 Complete: Cross-encoder ranking applied.")
        
        return candidates_df
    
    except Exception as e:
        st.error(f"Error in Stage 2 Contextual Re-ranking: {str(e)}")
        raise

def integrate_signals_and_score(
    candidates_df: pd.DataFrame,
    semantic_weight: float,
    ranking_weight: float,
    status_container
) -> pd.DataFrame:
    """Integrate RedRob behavioral signals and compute final scores."""
    with status_container.container():
        st.info("🔗 Integrating RedRob behavioral signals...")
    
    try:
        # Normalize scores to [0, 1] range
        if candidates_df["semantic_similarity"].max() > 0:
            candidates_df["semantic_norm"] = (
                candidates_df["semantic_similarity"] / candidates_df["semantic_similarity"].max()
            )
        else:
            candidates_df["semantic_norm"] = 0
        
        if candidates_df["cross_encoder_score"].max() > candidates_df["cross_encoder_score"].min():
            candidates_df["cross_encoder_norm"] = (
                (candidates_df["cross_encoder_score"] - candidates_df["cross_encoder_score"].min()) /
                (candidates_df["cross_encoder_score"].max() - candidates_df["cross_encoder_score"].min())
            )
        else:
            candidates_df["cross_encoder_norm"] = 0
        
        # Extract RedRob signals
        candidates_df["profile_completeness"] = candidates_df["profile"].apply(
            lambda x: x.get("profile_completeness_score", 0) / 100 if isinstance(x, dict) else 0
        )
        
        candidates_df["redrob_engagement"] = candidates_df["redrob_signals"].apply(
            lambda x: _compute_engagement_score(x) if isinstance(x, dict) else 0
        )
        
        # Compute composite final score
        candidates_df["ai_fit_score"] = (
            (semantic_weight * candidates_df["semantic_norm"]) +
            (ranking_weight * candidates_df["cross_encoder_norm"])
        ) * 100  # Scale to 0-100
        
        candidates_df["ai_fit_score"] = candidates_df["ai_fit_score"].round(2)
        
        with status_container.container():
            st.success(f"✅ Signal integration complete: Final scores computed.")
        
        return candidates_df
    
    except Exception as e:
        st.error(f"Error integrating signals: {str(e)}")
        raise

def _compute_engagement_score(redrob_dict: Dict) -> float:
    """Compute engagement score from RedRob signals."""
    if not isinstance(redrob_dict, dict):
        return 0.0
    
    weights = {
        "profile_completeness_score": 0.2,
        "open_to_work_flag": 0.15,
        "recruiter_response_rate": 0.15,
        "connection_count": 0.1,
        "endorsements_received": 0.1,
        "interview_completion_rate": 0.1,
        "github_activity_score": 0.1,
        "verified_email": 0.1
    }
    
    score = 0.0
    
    if "profile_completeness_score" in redrob_dict:
        score += (redrob_dict["profile_completeness_score"] / 100) * weights["profile_completeness_score"]
    
    if redrob_dict.get("open_to_work_flag", False):
        score += weights["open_to_work_flag"]
    
    if "recruiter_response_rate" in redrob_dict:
        score += redrob_dict["recruiter_response_rate"] * weights["recruiter_response_rate"]
    
    if "connection_count" in redrob_dict:
        conn_normalized = min(redrob_dict["connection_count"] / 500, 1.0)
        score += conn_normalized * weights["connection_count"]
    
    if "endorsements_received" in redrob_dict:
        endorse_normalized = min(redrob_dict["endorsements_received"] / 100, 1.0)
        score += endorse_normalized * weights["endorsements_received"]
    
    if "interview_completion_rate" in redrob_dict:
        score += redrob_dict.get("interview_completion_rate", 0) * weights["interview_completion_rate"]
    
    github_score = redrob_dict.get("github_activity_score", -1)
    if github_score >= 0:
        github_normalized = min(github_score / 100, 1.0)
        score += github_normalized * weights["github_activity_score"]
    else:
        score += 0
    
    if redrob_dict.get("verified_email", False):
        score += weights["verified_email"]
    
    return min(score, 1.0)

def generate_ai_reasoning(row: pd.Series, config: Dict) -> str:
    """Generate AI reasoning for each candidate."""
    candidate_id = row.get("candidate_id", "N/A")
    semantic_sim = row.get("semantic_similarity", 0)
    cross_encoder = row.get("cross_encoder_score", 0)
    years_exp = row.get("profile", {}).get("years_of_experience", 0) if isinstance(row.get("profile"), dict) else 0
    
    reasoning = f"Candidate {candidate_id}: "
    
    if semantic_sim >= 0.6:
        reasoning += "Strong semantic alignment with JD. "
    elif semantic_sim >= 0.4:
        reasoning += "Moderate semantic alignment. "
    else:
        reasoning += "Low semantic alignment. "
    
    if cross_encoder >= 0.5:
        reasoning += "High contextual relevance. "
    elif cross_encoder >= 0.3:
        reasoning += "Moderate contextual relevance. "
    else:
        reasoning += "Lower contextual relevance. "
    
    if years_exp >= 5:
        reasoning += f"Experienced professional ({years_exp:.1f} years). "
    else:
        reasoning += f"Early-career candidate ({years_exp:.1f} years). "
    
    return reasoning.strip()

# ============================================================================
# EXCEL EXPORT
# ============================================================================
def export_to_excel(results_df: pd.DataFrame, filename: str = "final_submission.xlsx") -> bytes:
    """Export results to Excel with formatting."""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Rankings"
        
        export_columns = ["Rank", "Candidate ID", "AI Fit Score", "AI Reasoning"]
        
        for col_num, col_title in enumerate(export_columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = col_title
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F77E8", end_color="1F77E8", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for row_num, (_, row) in enumerate(results_df.iterrows(), 2):
            ws.cell(row=row_num, column=1).value = row_num - 1
            ws.cell(row=row_num, column=2).value = row.get("candidate_id", "")
            ws.cell(row=row_num, column=3).value = row.get("ai_fit_score", 0)
            ws.cell(row=row_num, column=4).value = row.get("ai_reasoning", "")
        
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 60
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    except Exception as e:
        st.error(f"Error exporting to Excel: {str(e)}")
        return None

# ============================================================================
# STREAMLIT UI
# ============================================================================
def main():
    # Load configuration
    config = load_config("config.yaml")
    
    # ========== HEADER ==========
    st.markdown("""
    <h1 style='text-align: center; color: #58A6FF;'>🚀 SMASHERS AI Hiring Copilot</h1>
    <p style='text-align: center; color: #C9D1D9; font-size: 16px;'>
    Two-Stage Hybrid AI Ranking Pipeline: Semantic Filtering + Contextual Re-Ranking
    </p>
    """, unsafe_allow_html=True)
    
    st.markdown("---")
    
    # ========== SIDEBAR: CONFIGURATION ==========
    with st.sidebar:
        st.markdown("### ⚙️ Pipeline Configuration")

        st.markdown("**Adjustable Parameters:**")

        top_candidates = st.slider(
            "Top Candidates to Return",
            min_value=10,
            max_value=500,
            value=int(config.get("top_candidates", 100)),
            step=10,
            help="Number of top candidates to extract after semantic filtering"
        )

        semantic_weight = st.slider(
            "Semantic Weight",
            min_value=0.0,
            max_value=1.0,
            value=float(config.get("semantic_weight", 0.3)),
            step=0.1,
            help="Weight for semantic similarity in final score"
        )

        ranking_weight = st.slider(
            "Ranking Weight",
            min_value=0.0,
            max_value=1.0,
            value=float(config.get("ranking_weight", 0.7)),
            step=0.1,
            help="Weight for cross-encoder ranking in final score"
        )

        similarity_threshold = st.slider(
            "Similarity Threshold",
            min_value=0.0,
            max_value=1.0,
            value=float(config.get("similarity_threshold", 0.3)),
            step=0.05,
            help="Minimum semantic similarity to pass Stage 1"
        )
        
        # Normalize weights
        total_weight = semantic_weight + ranking_weight
        if total_weight > 0:
            semantic_weight = semantic_weight / total_weight
            ranking_weight = ranking_weight / total_weight
        
        st.divider()
        
        device = get_device()
        st.markdown(f"**System Info:**")
        st.caption(f"🖥️ Device: {str(device).upper()}")
        if str(device) == "cuda":
            st.caption(f"✅ GPU Acceleration: Enabled")
        else:
            st.caption(f"⚠️ GPU Acceleration: Disabled (CPU mode)")
    
    # ========== MAIN CONTENT: FILE UPLOADS ==========
    st.markdown("### 📂 Upload Your Data")

    input_mode = st.radio(
        "Select input method:",
        ["Upload Files", "Use Local File Path"],
        horizontal=True,
        label_visibility="collapsed"
    )

    jd_file = None
    candidates_file = None
    jd_path = None
    candidates_path = None

    if input_mode == "Upload Files":
        col_jd, col_candidates = st.columns(2)

        with col_jd:
            st.markdown("**Job Description (.docx or .pdf)**")
            jd_file = st.file_uploader(
                "Upload Job Description",
                type=["docx", "pdf"],
                key="jd_upload",
                label_visibility="collapsed"
            )

        with col_candidates:
            st.markdown("**Candidate Pool (.jsonl)**")
            candidates_file = st.file_uploader(
                "Upload Candidate Pool",
                type=["jsonl"],
                key="candidates_upload",
                label_visibility="collapsed"
            )
    else:
        col_jd, col_candidates = st.columns(2)

        with col_jd:
            st.markdown("**Job Description File Path**")
            jd_path = st.text_input(
                "Enter full path to Job Description (.docx or .pdf)",
                placeholder="C:\\path\\to\\job_description.docx",
                label_visibility="collapsed"
            )

        with col_candidates:
            st.markdown("**Candidate Pool File Path**")
            candidates_path = st.text_input(
                "Enter full path to Candidate Pool (.jsonl)",
                placeholder="C:\\path\\to\\candidates.jsonl",
                label_visibility="collapsed"
            )
    
    # ========== EXECUTION BUTTON & PIPELINE ==========
    st.markdown("---")

    # Determine if files are ready
    files_ready = False
    if input_mode == "Upload Files":
        files_ready = jd_file is not None and candidates_file is not None
    else:
        files_ready = jd_path and candidates_path and os.path.exists(jd_path) and os.path.exists(candidates_path)

    col_run, col_status = st.columns([1, 3])

    with col_run:
        run_button = st.button(
            "▶️ Run AI Copilot Ranking Pipeline",
            disabled=not files_ready,
            use_container_width=True,
            type="primary"
        )

    with col_status:
        if not files_ready:
            if input_mode == "Upload Files":
                st.warning("⏳ Upload both files to enable pipeline execution")
            else:
                if not jd_path or not os.path.exists(jd_path):
                    st.warning("⏳ Invalid Job Description file path")
                elif not candidates_path or not os.path.exists(candidates_path):
                    st.warning("⏳ Invalid Candidate Pool file path")
                else:
                    st.warning("⏳ Check file paths")
    
    # ========== PIPELINE EXECUTION ==========
    if run_button:
        status_placeholder = st.empty()

        try:
            with status_placeholder.container():
                st.info("📥 Parsing input files...")

            # Get files based on input mode
            if input_mode == "Upload Files":
                jd_file_to_parse = jd_file
                candidates_file_to_parse = candidates_file
            else:
                # Open files from disk paths
                jd_file_to_parse = open(jd_path, 'rb')
                candidates_file_to_parse = open(candidates_path, 'r')

            # Parse JD
            jd_text = parse_job_description(jd_file_to_parse)

            # Parse candidates
            candidates_df = parse_candidate_pool(candidates_file_to_parse)
            
            with status_placeholder.container():
                st.success(f"✅ Parsed {len(candidates_df)} candidates from pool.")
            
            # Load models
            with status_placeholder.container():
                st.info("🤖 Loading AI models (this may take a moment)...")
            
            semantic_model = load_semantic_model(config.get("model_semantic"))
            ranking_model = load_ranking_model(config.get("model_ranking"))
            
            with status_placeholder.container():
                st.success("✅ Models loaded successfully.")
            
            # ===== STAGE 1: SEMANTIC FILTERING =====
            st.markdown("### 🔄 Pipeline Execution Progress")
            
            progress_container = st.container()
            
            with progress_container:
                stage1_status = st.status("Stage 1: Semantic Filtering", expanded=True)
            
            filtered_candidates, embeddings_dict = stage_1_semantic_filtering(
                jd_text,
                candidates_df.copy(),
                semantic_model,
                similarity_threshold,
                top_candidates,
                stage1_status
            )
            
            stage1_status.update(label="Stage 1: Semantic Filtering ✅", state="complete")
            
            # ===== STAGE 2: CONTEXTUAL RE-RANKING =====
            with progress_container:
                stage2_status = st.status("Stage 2: Contextual Re-Ranking", expanded=True)
            
            ranked_candidates = stage_2_contextual_reranking(
                jd_text,
                filtered_candidates.copy(),
                ranking_model,
                stage2_status
            )
            
            stage2_status.update(label="Stage 2: Contextual Re-Ranking ✅", state="complete")
            
            # ===== SIGNAL INTEGRATION & SCORING =====
            with progress_container:
                stage3_status = st.status("Stage 3: Signal Integration & Scoring", expanded=True)
            
            final_results = integrate_signals_and_score(
                ranked_candidates.copy(),
                semantic_weight,
                ranking_weight,
                stage3_status
            )
            
            final_results["ai_reasoning"] = final_results.apply(
                lambda row: generate_ai_reasoning(row, config), axis=1
            )
            
            final_results = final_results.sort_values("ai_fit_score", ascending=False)
            final_results.reset_index(drop=True, inplace=True)
            
            stage3_status.update(label="Stage 3: Signal Integration & Scoring ✅", state="complete")
            
            st.markdown("---")
            
            # ========== RESULTS DISPLAY ==========
            st.markdown("### 📊 Top Candidates Rankings")
            
            display_df = pd.DataFrame({
                "Rank": range(1, len(final_results) + 1),
                "Candidate ID": final_results["candidate_id"],
                "AI Fit Score": final_results["ai_fit_score"],
                "AI Reasoning": final_results["ai_reasoning"]
            })
            
            # Display with progress bars for scores
            st.dataframe(
                display_df,
                use_container_width=True,
                height=500,
                column_config={
                    "AI Fit Score": st.column_config.ProgressColumn(
                        "AI Fit Score",
                        min_value=0,
                        max_value=100,
                        format="%.2f"
                    )
                }
            )
            
            st.markdown("---")
            
            # ========== EXPORT SECTION ==========
            st.markdown("### 💾 Export Results")
            
            col_export, col_info = st.columns([1, 2])
            
            with col_export:
                excel_data = export_to_excel(display_df)
                if excel_data:
                    st.download_button(
                        label="📥 Download as Excel (final_submission.xlsx)",
                        data=excel_data,
                        file_name="final_submission.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        type="primary"
                    )
            
            with col_info:
                st.metric("Total Candidates Ranked", len(display_df))
                st.metric("Top Candidate Score", f"{display_df['AI Fit Score'].max():.2f}")
            
            # ========== DETAILED CANDIDATE INFO ==========
            st.markdown("---")
            st.markdown("### 🔍 Detailed Candidate Information")
            
            selected_rank = st.selectbox(
                "Select a candidate to view details:",
                range(1, len(display_df) + 1),
                format_func=lambda x: f"Rank {x}: {display_df.iloc[x-1]['Candidate ID']}"
            )
            
            if selected_rank:
                candidate_idx = selected_rank - 1
                candidate_data = final_results.iloc[candidate_idx]
                
                detail_col1, detail_col2 = st.columns(2)
                
                with detail_col1:
                    st.markdown(f"**Candidate ID:** {candidate_data.get('candidate_id', 'N/A')}")
                    profile = candidate_data.get("profile", {})
                    if isinstance(profile, dict):
                        st.markdown(f"**Name:** {profile.get('anonymized_name', 'N/A')}")
                        st.markdown(f"**Headline:** {profile.get('headline', 'N/A')}")
                        st.markdown(f"**Experience:** {profile.get('years_of_experience', 'N/A')} years")
                        st.markdown(f"**Current Role:** {profile.get('current_title', 'N/A')} @ {profile.get('current_company', 'N/A')}")
                
                with detail_col2:
                    st.markdown(f"**AI Fit Score:** {candidate_data.get('ai_fit_score', 0):.2f}/100")
                    st.markdown(f"**Semantic Similarity:** {candidate_data.get('semantic_similarity', 0):.4f}")
                    st.markdown(f"**Cross-Encoder Score:** {candidate_data.get('cross_encoder_score', 0):.4f}")
                    st.markdown(f"**AI Reasoning:** {candidate_data.get('ai_reasoning', 'N/A')}")
                
                if isinstance(profile, dict) and "summary" in profile:
                    st.markdown("**Professional Summary:**")
                    st.markdown(profile.get("summary", "No summary available"))
        
        except Exception as e:
            st.error(f"❌ Pipeline execution failed: {str(e)}")
            st.exception(e)

if __name__ == "__main__":
    main()
